import openpyxl

from modules.trainee import trainee


def create_trainee():
    id = input("Enter trainee ID: ")
    name = input("Enter trainee name: ")
    course = input("Enter course: ")
    degree = input("Enter degree: ")
    work_exp = input("Enter work experience: ")
    t = trainee(id, name, course, degree, work_exp)
    t.save_to_excel()


def delete_trainee():
    id = input("Enter trainee ID: ")
    t = trainee(id, '', '', '', '')
    t.delete_from_excel()


def view_trainee():
    id = input("Enter trainee ID: ")
    wb = openpyxl.load_workbook('MasterRecord.xlsx')
    ws = wb['ListOfTrainees']
    for row in ws.iter_rows(min_row=2, max_col=5):
        if row[0].value == id:
            print(f"ID: {row[0].value}")
            print(f"Name: {row[1].value}")
            print(f"Course: {row[2].value}")
            print(f"Degree: {row[3].value}")
            print(f"Work Experience: {row[4].value}")
            break
    else:
        print("Trainee not found.")


def view_all_trainees():
    wb = openpyxl.load_workbook('MasterRecord.xlsx')
    ws = wb['ListOfTrainees']
    for row in ws.iter_rows(min_row=2, max_col=5):
        print(f"ID: {row[0].value}")
        print(f"Name: {row[1].value}")
        print(f"Course: {row[2].value}")
        print(f"Degree: {row[3].value}")
        print(f"Work Experience: {row[4].value}")
        print()


while True:
    print("1. Create trainee")
    print("2. Delete trainee")
    print("3. View trainee")
    print("4. View all trainees")
    print("5. Exit")

    choice = input("Enter your choice: ")

    if choice == '1':
        create_trainee()
    elif choice == '2':
        delete_trainee()
    elif choice == '3':
        view_trainee()
    elif choice == '4':
        view_all_trainees()
    elif choice == '5':
        break
    else:
        print("Invalid choice. Please try again.")
