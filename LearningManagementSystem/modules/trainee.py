import openpyxl

class trainee:
    def __init__(self, id, name, course, degree, work_exp):
        self._id = id
        self._name = name
        self._course = course
        self._degree = degree
        self._work_exp = work_exp

    def get_id(self):
        return self._id

    def set_id(self, id):
        self._id = id

    def get_name(self):
        return self._name

    def set_name(self, name):
        self._name = name

    def get_course(self):
        return self._course

    def set_course(self, course):
        self._course = course

    def get_degree(self):
        return self._degree

    def set_degree(self, degree):
        self._degree = degree

    def get_work_exp(self):
        return self._work_exp

    def set_work_exp(self, work_exp):
        self._work_exp = work_exp

    def save_to_excel(self):
        wb = openpyxl.load_workbook('MasterRecord.xlsx')
        ws = wb['ListOfTrainees']
        last_row = ws.max_row
        ws.cell(row=last_row+1, column=1, value=self._id)
        ws.cell(row=last_row+1, column=2, value=self._name)
        ws.cell(row=last_row+1, column=3, value=self._course)
        ws.cell(row=last_row+1, column=4, value=self._degree)
        ws.cell(row=last_row+1, column=5, value=self._work_exp)
        wb.save('MasterRecord.xlsx')
